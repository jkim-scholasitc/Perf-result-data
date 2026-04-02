#!/usr/bin/env python3
"""
BlazeMeter CSV to Excel Converter

This script converts BlazeMeter test result CSV files to Performance Test Result Template format.
It extracts summary data from the "ALL" row and creates a formatted report.

Usage:
    python blazemeter_to_excel.py <input_csv_file> <test_type> [output_xlsx_file]
    
    test_type: API or UI
    - API: P95% > 500ms = Fail (red), P95% <= 500ms = Pass (blue)
    - UI: P95% > 2000ms = Fail (red), P95% <= 2000ms = Pass (blue)

Example:
    python blazemeter_to_excel.py data.csv API
    # Creates: data-API-converted-YYYYMMDD-HHMMSS.xlsx
    
    python blazemeter_to_excel.py data.csv UI output.xlsx
    # Creates: output.xlsx
"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil


def convert_blazemeter_to_excel(input_csv, test_type='API', output_xlsx=None):
    """
    Convert BlazeMeter CSV to Performance Test Result Template format.
    
    Args:
        input_csv (str): Path to input CSV file
        test_type (str): Test type - 'API' or 'UI'
        output_xlsx (str, optional): Path to output Excel file. 
                                     If None, uses input filename with timestamp
    
    Returns:
        str: Path to the created Excel file
    """
    # Validate input file exists
    if not os.path.exists(input_csv):
        raise FileNotFoundError(f"Input file not found: {input_csv}")
    
    # Determine output filename with test type and timestamp
    if output_xlsx is None:
        base_name = os.path.splitext(input_csv)[0]
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        output_xlsx = f"{base_name}-{test_type.upper()}-converted-{timestamp}.xlsx"
    
    print(f"\nReading CSV file: {input_csv}")
    
    # Read the CSV file
    df = pd.read_csv(input_csv)
    
    # Extract the "ALL" row data for summary
    all_row = df[df['Element Label'] == 'ALL']
    
    if all_row.empty:
        raise ValueError("No 'ALL' row found in the CSV file. Cannot extract summary data.")
    
    # Extract summary values from ALL row
    avg_bandwidth = all_row['Avg. Bandwidth (KBytes/s)'].values[0]
    total_hits = int(all_row['# Samples'].values[0])
    avg_hits_per_sec = all_row['Avg. Hits/s'].values[0]
    response_95 = int(all_row['95% line (ms)'].values[0])
    errors = all_row['Error Percentage'].values[0] / 100  # Convert to decimal
    avg_response_time = all_row['Avg. Response Time (ms)'].values[0]
    
    print(f"  ✓ Extracted summary data from 'ALL' row")
    
    # Remove the ALL row from transaction data
    df_transactions = df[df['Element Label'] != 'ALL'].copy()
    
    # Sort by Element Label in ascending order
    df_transactions = df_transactions.sort_values(by='Element Label', ascending=True)
    print(f"  ✓ Sorted transaction data by 'Element Label' (ascending)")
    
    # Define the desired column order for transactions
    column_order = [
        'Element Label',
        '# Samples',
        'Avg. Response Time (ms)',
        '95% line (ms)',
        'Error Percentage',
        'Avg. Hits/s',
        '90% line (ms)',
        '99% line (ms)',
        'Min Response Time (ms)',
        'Max Response Time (ms)',
        'Avg. Bandwidth (KBytes/s)',
        'Concurrency'
    ]
    
    # Reorder columns
    existing_columns = [col for col in column_order if col in df_transactions.columns]
    df_transactions = df_transactions[existing_columns]
    
    # Check if template exists
    template_path = '/Users/jameskim/Documents/Scripts/Python/Performance-Test-result-Template.xlsx'
    
    if os.path.exists(template_path):
        print(f"  ✓ Using template: {template_path}")
        # Copy template to output location
        shutil.copy(template_path, output_xlsx)
        workbook = load_workbook(output_xlsx)
        worksheet = workbook['Performance Test Report']
        
        # Unmerge all cells in the worksheet to allow writing
        merged_cells_to_unmerge = list(worksheet.merged_cells)
        
        for merged_cell in merged_cells_to_unmerge:
            worksheet.unmerge_cells(str(merged_cell))
        
        print(f"  ✓ Unmerged all cells for data update")
        
        # Update Row 1: Test name (should be the input filename without extension)
        test_name = os.path.splitext(os.path.basename(input_csv))[0]
        worksheet['B1'] = test_name
        
        # Re-merge Row 1 cells B1:F1 after writing
        worksheet.merge_cells('B1:F1')
        
        # Update Row 2: Date (should be empty)
        worksheet['B2'] = ''
        
        # Re-merge Row 2 cells B2:F2 after clearing
        worksheet.merge_cells('B2:F2')
        
        print(f"  ✓ Updated Test name (Row 1) with filename: {test_name}")
        print(f"  ✓ Cleared Date field (Row 2)")
        
        # Update summary data in the template (columns B-F, rows 3-7)
        # Row 3: Average Throughput
        worksheet['B3'] = avg_bandwidth
        worksheet.merge_cells('B3:F3')
        
        # Row 4: Total Hits
        worksheet['B4'] = total_hits
        worksheet.merge_cells('B4:F4')
        
        # Row 5: Average Hits per Second
        worksheet['B5'] = avg_hits_per_sec
        worksheet.merge_cells('B5:F5')
        
        # Row 6: 95% Response Time
        response_time_str = f"{int(avg_response_time)}/{response_95}"
        worksheet['B6'] = response_time_str
        worksheet.merge_cells('B6:F6')
        
        # Row 7: Errors
        worksheet['B7'] = errors
        worksheet.merge_cells('B7:F7')
        
        print(f"  ✓ Updated summary data in template")
        
        # Update Failed count of Transaction (row 10) and Passed count of Transaction (row 11)
        # These will be calculated after processing all transactions
        
        # Clear existing transaction data (starting from row 16)
        max_row = worksheet.max_row
        if max_row > 15:
            worksheet.delete_rows(16, max_row - 15)
        
        # Add transaction data starting at row 16
        start_row = 16
        
        # Define styles for transaction rows
        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Write transaction data and count pass/fail
        failed_count = 0
        passed_count = 0
        
        for i, (idx, row) in enumerate(df_transactions.iterrows()):
            row_num = start_row + i
            
            # Column A: Transaction Name
            cell = worksheet.cell(row=row_num, column=1, value=row['Element Label'])
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Column B: Count
            cell = worksheet.cell(row=row_num, column=2, value=int(row['# Samples']))
            cell.alignment = number_alignment
            cell.border = thin_border
            
            # Column C: Avg
            cell = worksheet.cell(row=row_num, column=3, value=int(row['Avg. Response Time (ms)']))
            cell.alignment = number_alignment
            cell.border = thin_border
            
            # Column D: P95%
            cell = worksheet.cell(row=row_num, column=4, value=int(row['95% line (ms)']))
            cell.alignment = number_alignment
            cell.border = thin_border
            
            # Column E: Error%
            cell = worksheet.cell(row=row_num, column=5, value=row['Error Percentage'])
            cell.alignment = number_alignment
            cell.border = thin_border
            
            # Column F: Result (pass/Fail based on test type and thresholds)
            p95_value = int(row['95% line (ms)'])
            error_pct = row['Error Percentage']
            
            if test_type.upper() == 'API':
                # API: P95% > 500ms = Fail, otherwise Pass
                if p95_value > 500:
                    result = 'Fail'
                    result_color = 'FF0000'  # Red
                    failed_count += 1
                else:
                    result = 'Pass'
                    result_color = '0000FF'  # Blue
                    passed_count += 1
            else:
                # UI: P95% > 2000ms = Fail, otherwise Pass
                if p95_value > 2000:
                    result = 'Fail'
                    result_color = 'FF0000'  # Red
                    failed_count += 1
                else:
                    result = 'Pass'
                    result_color = '0000FF'  # Blue
                    passed_count += 1
            
            cell = worksheet.cell(row=row_num, column=6, value=result)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(color=result_color, bold=True)
        
        # Add ALL row at the end
        all_row_num = start_row + len(df_transactions)
        worksheet.cell(row=all_row_num, column=1, value='ALL')
        worksheet.cell(row=all_row_num, column=2, value=total_hits)
        worksheet.cell(row=all_row_num, column=3, value=int(avg_response_time))
        worksheet.cell(row=all_row_num, column=4, value=response_95)
        worksheet.cell(row=all_row_num, column=5, value=errors * 100)
        # Determine ALL row result
        if test_type.upper() == 'API':
            all_result = 'Fail' if response_95 > 500 else 'Pass'
            all_result_color = 'FF0000' if response_95 > 500 else '0000FF'
        else:
            all_result = 'Fail' if response_95 > 2000 else 'Pass'
            all_result_color = 'FF0000' if response_95 > 2000 else '0000FF'
        
        worksheet.cell(row=all_row_num, column=6, value=all_result)
        
        # Apply formatting to ALL row
        for col in range(1, 7):
            cell = worksheet.cell(row=all_row_num, column=col)
            cell.border = thin_border
            if col == 6:
                cell.font = Font(color=all_result_color, bold=True)
            if col > 1:
                cell.alignment = number_alignment
            else:
                cell.alignment = cell_alignment
        
        # Update row 10: Failed count of Transaction
        worksheet['B10'] = failed_count
        worksheet.merge_cells('B10:F10')
        
        # Update row 11: Passed count of Transaction
        worksheet['B11'] = passed_count
        worksheet.merge_cells('B11:F11')
        
        # Update row 12: Result (Fail if any of these conditions are true)
        # 1. Error rate > 1%
        # 2. P95% > threshold (500ms for API, 2000ms for UI)
        # 3. Any failed transactions exist
        
        error_rate_fail = (errors * 100) > 1
        
        if test_type.upper() == 'API':
            p95_fail = response_95 > 500
        else:
            p95_fail = response_95 > 2000
        
        any_transaction_failed = failed_count > 0
        
        # Overall result is Fail if ANY condition is true
        overall_result = 'Fail' if (error_rate_fail or p95_fail or any_transaction_failed) else 'Pass'
        overall_result_color = 'FF0000' if (error_rate_fail or p95_fail or any_transaction_failed) else '0000FF'
        
        worksheet['B12'] = overall_result
        worksheet['B12'].font = Font(color=overall_result_color, bold=True)
        worksheet.merge_cells('B12:F12')
        
        # Update row 13: Analysis - explain why test failed or passed
        analysis_parts = []
        
        if error_rate_fail:
            analysis_parts.append(f"Error rate is high ({errors*100:.2f}%)")
        
        if any_transaction_failed:
            analysis_parts.append(f"There are {failed_count} failed transaction(s)")
        
        if p95_fail:
            sla_threshold = 500 if test_type.upper() == 'API' else 2000
            analysis_parts.append(f"Overall P95% response time is over SLA ({response_95} ms > {sla_threshold} ms)")
        
        if analysis_parts:
            analysis_text = ". ".join(analysis_parts) + "."
        else:
            analysis_text = "All metrics are within acceptable thresholds."
        
        # Write analysis to row 13
        worksheet['B13'] = analysis_text
        worksheet.merge_cells('B13:F13')
        
        # Row 14: System Stats (keep empty but merge cells)
        worksheet['B14'] = ''
        worksheet.merge_cells('B14:F14')
        
        print(f"  ✓ Added {len(df_transactions)} transactions to report")
        print(f"  ✓ Failed transactions: {failed_count}")
        print(f"  ✓ Passed transactions: {passed_count}")
        print(f"  ✓ Overall result: {overall_result}")
        print(f"  ✓ Analysis: {analysis_text}")
        
    else:
        # Create new workbook without template
        print(f"  ⚠ Template not found, creating basic report")
        
        # Create a new Excel file with summary and transaction data
        with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
            # Create summary sheet
            summary_data = {
                'Metric': [
                    'Average Throughput (Kbytes/second)',
                    'Total Hits',
                    'Average Hits per Second(TPS)',
                    '95% Response Time(ms)',
                    'Errors'
                ],
                'Value': [
                    avg_bandwidth,
                    total_hits,
                    avg_hits_per_sec,
                    f"{int(avg_response_time)}/{response_95}",
                    errors
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Performance Test Report', index=False, startrow=2)
            
            # Add transaction data
            df_transactions.to_excel(writer, sheet_name='Performance Test Report', index=False, startrow=10)
        
        # Apply formatting
        workbook = load_workbook(output_xlsx)
        worksheet = workbook['Performance Test Report']
        
        # Format headers
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            
            cell = worksheet.cell(row=11, column=col)
            cell.font = header_font
            cell.fill = header_fill
    
    # Save the workbook
    workbook.save(output_xlsx)
    print(f"  ✓ Successfully created: {output_xlsx}")
    
    # Print summary
    print(f"\n  Summary:")
    print(f"    - Total transactions: {len(df_transactions)}")
    print(f"    - Total hits: {total_hits:,}")
    print(f"    - Average throughput: {avg_bandwidth} KB/s")
    print(f"    - Error rate: {errors*100:.2f}%")
    print(f"    - Output file size: {os.path.getsize(output_xlsx):,} bytes")
    
    return output_xlsx


def main():
    """Main function to handle command-line execution."""
    if len(sys.argv) < 3:
        print(__doc__)
        print("\nError: Please provide an input CSV file and test type (API or UI).")
        print("\nUsage:")
        print("  python blazemeter_to_excel.py <input_csv_file> <test_type> [output_xlsx_file]")
        print("\nTest Type:")
        print("  API - P95% > 500ms = Fail (red), P95% <= 500ms = Pass (blue)")
        print("  UI  - P95% > 2000ms = Fail (red), P95% <= 2000ms = Pass (blue)")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    test_type = sys.argv[2]
    output_xlsx = sys.argv[3] if len(sys.argv) > 3 else None
    
    # Validate test type
    if test_type.upper() not in ['API', 'UI']:
        print(f"\n✗ Error: Invalid test type '{test_type}'. Must be 'API' or 'UI'.")
        sys.exit(1)
    
    try:
        convert_blazemeter_to_excel(input_csv, test_type, output_xlsx)
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
